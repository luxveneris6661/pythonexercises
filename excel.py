# goal: generate an excel file named 'sample' with 4 columns corresponding to dates, products, branches, and profit
# and randomly generated data in rows 2-500


#import all needed modules: openpyxl, date, timedelta, random
#define: generate worksheet, generate random date,  names of each column, data constants
#make sure the generated data is in date order
#generate rows 2-10 and assign random values to each column
#adjust column width
#save
#generate worksheet 


import openpyxl as px #'as' shorthand alias, referring to openpyxl as 'px'
from datetime import date, timedelta
import random 
import os

#Data Constants:

products = ('Product 1', 'Product 2')
branches = ('Branch 1', 'Branch 2')
start_date = date(2025,1,1)
end_date = date(2025,10,19)


#function to create random dates:
def random_date (start_date: date, end_date: date) :
    delta_days = (end_date - start_date).days
    return start_date + timedelta(days = random.randrange(delta_days))

#function to generate excel file:

def generate (filename: str):
 wb = px.Workbook() 
 ws = wb.active #gets the currently active worksheet
 ws.title = 'sheet_1'#names the worksheet

 #naming each column: 

 ws ['A1'] = 'Day'
 ws ['B1'] = 'Product'
 ws ['C1'] = 'Branch'
 ws ['D1'] = 'Profit'

 #making sure generated data is sorted in date order:

 data_rows = [] #collects data in empty list variable
 for _ in range(2,11): 
  row_date = random_date(start_date, end_date) 
  row_product = random.choice(products)
  row_branch = random.choice(branches)
  row_profit = random.randint(10000,100000)
  data_rows.append((row_date, row_product, row_branch, row_profit)) #tuple in the append() method, hence the double brackets *append accepts one argument

 data_rows.sort(key = lambda x: x[0]) #sort usually compares all tuples. key tells sort what to sort by. we only want to sort by x[0](first element of tuple, 'date'). lambda x = x[0] returns the first element of the tuple, which is our key.

 #randomly generating data *note that I defined the variable 'row' here:

 for i, row in enumerate (data_rows, start = 2):  #yields (index,element) pairs; 'i' -> excel row number; 'row' tuple is single row in data_row list 
 #“For each enumerated pair (index, element) in the sequence 'data_rows', assign the first part (index no.) to i, and the second part to row. Start indexing at 2.”
  ws[f'A{i}'] = row[0] # at cell A{i}, write the first element of the current iteration's 'row' tuple (the date). Shorthand for cell_obj = ws['A1'] \n cell_obj.value = 'Hello'
  ws[f'B{i}'] = row[1]
  ws[f'C{i}'] = row[2]
  ws[f'D{i}'] = row[3]

 #python assigns values of each cell, openpyxl writes them to actual excel file

 ws.column_dimensions['A'].width = 13 #adjusting column width

 
 #save workbook

 filepath = os.path.join(r'C:\Users\WURUOQING\Desktop\python', filename)

 wb.save(filepath)  

 os.startfile(filepath)


generate ('sample.xlsx')

