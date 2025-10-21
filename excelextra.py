#goal: generate excel sheet where sheet 1 "sum" is summized data and sheets 2-11 "No.1"-"No.10" are extracted data 

#import modules
#data constants
#def random date
#def generate:
#generate excel sheet with randomized info and headers as tables (enabling filtering in excel)
#write info into new indexed sheets

import openpyxl as px
import os
from datetime import date, timedelta
import random

products = ('Product 1', 'Product 2')
branches = ('Branch 1', 'Branch 2')
start_date = date(2025,1,1)
end_date = date(2025,10,20)
folder = (r"C:\Users\WURUOQING\Desktop\python")



def random_date (start_date:date, end_date:date):
    delta_days = (end_date - start_date).days
    return start_date + timedelta(days = random.randrange(delta_days))

def generate (filename: str):
    wb = px.Workbook()
    ws = wb.active
    ws.title = 'Sum'

    #naming each column: 

    ws ['A1'] = 'Day'
    ws ['B1'] = 'Product'
    ws ['C1'] = 'Branch'
    ws ['D1'] = 'Profit'
 
 #collects data in empty list variable
 #tuple in the append() method, hence the double brackets *append accepts one argument
    data_rows = []
    for _ in range (2,12):
        row_date = random_date(start_date, end_date)
        row_product = random.choice(products)
        row_branch = random.choice(branches)
        row_profit = random.randint(10000,100000)
        data_rows.append((row_date, row_product, row_branch, row_profit))

 
 #randomly generating data *note that I defined the variable 'row' here:
    for i, row in enumerate (data_rows, start = 2):
        ws[f'A{i}'] = row[0]
        ws[f'B{i}'] = row[1]
        ws[f'C{i}'] = row[2]
        ws[f'D{i}'] = row[3]

 #python assigns values of each cell, openpyxl writes them to actual excel file

 #adjusting column width

    ws.column_dimensions['A'].width = 13

 #making row 1 into tables
    ws.auto_filter.ref = 'A1:D11'
    

 #putting data from 'Sum' into disparate sheets

    for i, row in enumerate (ws.iter_rows(min_row = 2, max_row = 12, values_only = True), start = 2):
         new_ws = wb.create_sheet(title = f'No.{i-1}')
         new_ws.column_dimensions['A'].width = 13
         headers = ['Day', 'Product', 'Branch', 'Profit']

    #assigns each indexed header element to row 1 and column at index col (starting from 1) with value "header"
         for col, header in enumerate (headers, start = 1):
          new_ws.cell(row = 1, column = col, value = header)

    #assigns each iterated row element to row 2 and column at index col (starting from 1) with the value "value"
         for col, value in enumerate (row, start = 1):
          cell = new_ws.cell(row = 2, column = col, value = value)
          if col == 1:
             cell.number_format = 'yyyy-mm-dd'
            
 #save workbook
    filepath = os.path.join(folder,filename)
    wb.save(filepath)
    os.startfile(filepath)

generate('Sum.xlsx')

