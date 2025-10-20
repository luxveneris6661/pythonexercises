#goal: map data from rows 2-10 to 9 separate excel sheets

#import all needed modules: os, openpyxl,
#read data from sample.xlsx
#loop: write row n into samplen.xlsx

import os
import openpyxl as px

folder = r"C:\Users\WURUOQING\Desktop\python"
filename = "sample.xlsx"
filepath = os.path.join(folder, filename)
wb = px.load_workbook(filepath)  #openpyxl specific commands/operations
ws = wb.active

#for each row i, read data from i, write to new worksheet numbered i

for i, row in enumerate (ws.iter_rows(min_row = 2, max_row = 10, values_only=True), start = 2):#indexing the data from the original sheet
 new_ws = wb.create_sheet(title=f'Sheet_{i}') #creates a new worksheet with each iteration i

 headers = ['Day', 'Product', 'Branch', 'Profit'] #defines the values of the headers

 for col, header in enumerate (headers, start = 1): #assigns each indexed header element to row 1 and column at index col (starting from 1) with value "header"
  new_ws.cell(row = 1, column = col, value = header)
 for col, value in enumerate (row, start = 1): #assigns each iterated row element to row 2 and column at index col (starting from 1) with the va
  new_ws.cell(row = 2, column = col, value = value)
